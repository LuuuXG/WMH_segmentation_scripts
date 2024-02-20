#%%
import pandas as pd

# 加载文件
ref_adnimerge_df = pd.read_excel(r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023.xlsx')
conversion_info_df = pd.read_excel(r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\sourcedata\conversion_info_overall.xlsx')

# 创建映射字典
conversion_dict = dict(zip(conversion_info_df['Original Name'], conversion_info_df['Subject ID']))

# 为 REF_ADNIMERGE_26Nov2023-test.xlsx 添加 ID 列
ref_adnimerge_df['ID'] = ref_adnimerge_df['PTID'].map(conversion_dict)

# 保存修改后的文件
updated_file_path = r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023.xlsx'
ref_adnimerge_df.to_excel(updated_file_path, index=False)

print(f"Updated file saved at {updated_file_path}")


#%%

import pandas as pd
from datetime import datetime, timedelta

# Load the Excel files
modified_df = pd.read_excel(r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023.xlsx')
wmh_df = pd.read_excel(r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\sourcedata\WMH_indices_overall.xlsx')

# Function to parse the date from the ID in WMH_indices_overall
def parse_date_from_id(id_str):
    try:
        date_str = id_str.split('_')[-1]
        return datetime.strptime(date_str, "%Y%m%d")
    except:
        return None

# Apply the function to create a new column with parsed dates
wmh_df['Parsed_Date'] = wmh_df['ID'].apply(parse_date_from_id)

# 复制 wmh_df 中的 'ID' 列到一个新列 'WMH_ID'
wmh_df['WMH_ID'] = wmh_df['ID']

# Function to find the closest date in WMH data for each row in modified data
def find_closest_wmh(row, wmh_df, threshold=120):
    participant_wmh = wmh_df[wmh_df['ID'].str.contains(row['ID'])]
    min_diff = timedelta(days=threshold + 1)
    closest_idx = None
    for idx, wmh_row in participant_wmh.iterrows():
        diff = abs(row['EXAMDATE'] - wmh_row['Parsed_Date'])
        if diff < min_diff:
            min_diff = diff
            closest_idx = idx
    return closest_idx if min_diff <= timedelta(days=threshold) else None

# Applying the function to each row in the modified dataset
modified_df['Closest_WMH_Index'] = modified_df.apply(lambda row: find_closest_wmh(row, wmh_df), axis=1)

# Merging the relevant WMH data into the modified dataset
wmh_columns_to_merge = wmh_df.columns.difference(modified_df.columns).tolist()
wmh_reduced_df = wmh_df[wmh_columns_to_merge]
merged_df = pd.merge(modified_df, wmh_reduced_df, left_on='Closest_WMH_Index', right_index=True, how='left')

# Determine the final column order
# Keep the original order of columns in modified_df
final_columns_order = modified_df.columns.tolist()

# Append the columns from wmh_df that are not in modified_df, in the order they appear in wmh_df
for col in wmh_df.columns:
    if col not in modified_df.columns:
        final_columns_order.append(col)

# Reorder the columns in merged_df
merged_df = merged_df[final_columns_order]

# Saving the merged dataframe to a new Excel file
merged_file_path = r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023_full.xlsx'
merged_df.to_excel(merged_file_path, index=False)

print(f"Merged file saved at {merged_file_path}")


#%%

import pandas as pd

# Load the merged dataset
merged_df = pd.read_excel(r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023_full.xlsx')

# Dropping rows where 'ID_y' is missing
cleaned_df = merged_df.dropna(subset=['WMH_ID'])

# Saving the cleaned dataset to a new Excel file
cleaned_df.to_excel(r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023_full.xlsx', index=False)

#%%

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Load the datasets
merged_df_path = r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023_full.xlsx'
low_quality_df_path = r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\overall_low_quality_segmentation.xlsx'

merged_df = pd.read_excel(merged_df_path, engine='openpyxl')
low_quality_df = pd.read_excel(low_quality_df_path, engine='openpyxl')

# Extract IDs from low quality segmentation dataset
low_quality_ids = low_quality_df['ID'].tolist()

# Open the merged dataframe in openpyxl for formatting
wb = openpyxl.load_workbook(merged_df_path)
ws = wb.active

# Find the column index for 'WMH_ID'
wmh_id_col_index = None
for col in ws.iter_cols(min_row=1, max_row=1):
    for cell in col:
        if cell.value == 'WMH_ID':
            wmh_id_col_index = cell.col_idx
            break

if wmh_id_col_index is None:
    raise Exception("WMH_ID column not found in the merged dataset.")

# Define the red fill pattern
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# Iterate through the rows in the worksheet and apply red fill to rows with low quality IDs
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Skipping header row
    if row[wmh_id_col_index - 1].value in low_quality_ids:  # Adjusting for zero-based index
        for cell in row:
            cell.fill = red_fill

# Save the modified workbook
modified_merged_df_path = r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023_full_highlight.xlsx'
wb.save(modified_merged_df_path)

print(f"Modified file saved at {modified_merged_df_path}")

#%%

import pandas as pd

# Load the merged dataset
merged_df_path = r'E:\Projects\WMH_segmentation\data_analysis\data\sourcedata\overall\overall_data_for_analysis.xlsx'
merged_df = pd.read_excel(merged_df_path)

# Replace empty cells with 'NA'
merged_df.fillna('NA', inplace=True)

# Save the updated dataframe to a new Excel file
updated_merged_df_path = r'E:\Projects\WMH_segmentation\data_analysis\data\sourcedata\overall\overall_data_for_analysis.xlsx'
merged_df.to_excel(updated_merged_df_path, index=False)

print(f"Updated file saved at {updated_merged_df_path}")

#%%

import pandas as pd

# Load the Excel file
file_path = r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023_followup.xlsx'
data = pd.read_excel(file_path)

# Extracting the subject identifier and the time sequence number
data['Subject'] = data['WMH_ID'].apply(lambda x: x.split('_')[0].split('-')[1])
data['TimeSequence'] = data['WMH_ID'].apply(lambda x: int(x.split('_')[1]))

# Sorting and filtering the data
data_sorted = data.sort_values(by=['Subject', 'TimeSequence'])
baseline_data = data_sorted.drop_duplicates(subset='Subject', keep='first')

# Saving the filtered data to a new file
baseline_data.to_excel(r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023_followup.xlsx', index=False)

#%%

import pandas as pd

# Load the Excel file
file_path = r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023_include.xlsx'
data = pd.read_excel(file_path)

# Extracting the subject identifier from the 'WMH_ID' column
data['Subject'] = data['WMH_ID'].apply(lambda x: x.split('_')[0].split('-')[1])

# Count the number of records for each subject
subject_counts = data['Subject'].value_counts()

# Identifying subjects with exactly 1 record
subjects_with_single_record = subject_counts[subject_counts == 1].index

# Exclude these subjects from the original data
data_excluding_single_records = data[~data['Subject'].isin(subjects_with_single_record)]

# Sorting the data by 'WMH_ID'
sorted_data = data_excluding_single_records.sort_values(by='WMH_ID')

# Saving the sorted data to a new Excel file
sorted_file_path = r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023_followup.xlsx'
sorted_data.to_excel(sorted_file_path, index=False)

#%% 从ADNI结果添加其他数据

import pandas as pd

def load_data(file_path, sheet_name=0):
    """
    Load data from a specified Excel file and sheet name or index.
    """
    return pd.read_excel(file_path, sheet_name=sheet_name)

def match_dates_v2(ptid_group, reference_data, date_column, value_column):
    """
    Match each EXAMDATE in target_data with the closest and non-repeating date in reference_data.
    Handle cases where the number of dates don't match by filling unmatched dates with None.
    """
    ptid_ref_data = reference_data[reference_data['PTID'] == ptid_group.name]
    if ptid_ref_data.empty:
        # No matching PTID in reference data
        return pd.Series([None] * len(ptid_group), index=ptid_group.index)

    matched_dates = []
    for target_date in ptid_group['EXAMDATE']:
        # Calculate time differences and sort them
        time_diff = ptid_ref_data[date_column].apply(lambda date: abs(date - target_date))
        sorted_diff = time_diff.sort_values()

        for date_index in sorted_diff.index:
            if date_index not in matched_dates:
                matched_dates.append(date_index)
                break
        else:
            # No more unique dates to match, append None
            matched_dates.append(None)

    # Retrieve the value for the matched dates, handling cases where no match was found
    matched_values = [ptid_ref_data.loc[idx, value_column] if idx in ptid_ref_data.index and idx is not None else None
                      for idx in matched_dates]

    return pd.Series(matched_values, index=ptid_group.index)

def merge_data(target_file, target_sheet, reference_file, reference_sheet, date_column, value_column):
    """
    Merge data from reference file into target file based on PTID and closest date match.
    """
    # Load the datasets
    target_data = load_data(target_file, target_sheet)
    reference_data = load_data(reference_file, reference_sheet)

    # Apply the revised matching process to the target data
    matched_values = target_data.groupby('PTID').apply(lambda group: match_dates_v2(group, reference_data, date_column, value_column))
    matched_values = matched_values.reset_index(level=0, drop=True)

    # Add the matched values to the target dataframe
    target_data[f'Matched_{value_column}'] = matched_values

    return target_data
# Example usage
target_file = r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023_include_1.xlsx'
target_sheet = 0  # or the sheet name
reference_file = r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\toWZ_alc.xlsx'
reference_sheet = 1  # or the sheet name
date_column = 'VISDATE'  # Date column in the reference file
value_column = 'MH14ALCH'  # Column to be merged from the reference file

merged_data = merge_data(target_file, target_sheet, reference_file, reference_sheet, date_column, value_column)

# Save the final merged dataset to a new Excel file
final_output_file_path = r'F:\Public_dataset\ADNI_WMH_seg\data_analysis\REF_ADNIMERGE_26Nov2023_include_1.xlsx'
merged_data.to_excel(final_output_file_path, index=False)

