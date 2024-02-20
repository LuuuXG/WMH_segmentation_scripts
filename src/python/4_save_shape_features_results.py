import os
import pandas as pd

# 更新后的 calculate_means_ge 函数
def calculate_means_ge(data, column, thresholds):
    means = []
    for threshold in thresholds:
        filtered_data = data[data[column] >= threshold]
        mean_values = {col: (filtered_data[col].mean() if not filtered_data.empty else 'NA')
                       for col in data.columns if col != 'Region'}
        mean_values['Threshold'] = threshold  # 添加阈值信息
        means.append(mean_values)
    return means

# 指定目录
derivatives_dir = r'E:\Projects\WMH_segmentation\derivatives'

# 获取所有子目录
sub_dirs = [d for d in os.listdir(derivatives_dir) if
            os.path.isdir(os.path.join(derivatives_dir, d)) and d.startswith('sub-')]

# 阈值列表
thresholds = [5, 10, 50, 100, 200, 500]

# 遍历每个子目录
for sub_dir in sub_dirs:
    print('processing directory: {0}'.format(sub_dir))
    shape_features_dir = os.path.join(derivatives_dir, sub_dir, 'shape_features')

    # 构建文件路径
    PWMH_shape_features_path = os.path.join(shape_features_dir, f'{sub_dir}_PWMH_shape_features_5voxels.xlsx')
    DWMH_shape_features_path = os.path.join(shape_features_dir, f'{sub_dir}_DWMH_shape_features_5voxels.xlsx')

    # 读取数据
    PWMH_data = pd.read_excel(PWMH_shape_features_path)
    DWMH_data = pd.read_excel(DWMH_shape_features_path)

    # 移除'Region'列
    PWMH_data.drop(columns=['Region'], inplace=True)
    DWMH_data.drop(columns=['Region'], inplace=True)

    # 计算阈值
    PWMH_mean_values = calculate_means_ge(PWMH_data, 'Volume', thresholds)
    DWMH_mean_values = calculate_means_ge(DWMH_data, 'Volume', thresholds)

    # 转换结果为DataFrame
    PWMH_mean_values_df = pd.DataFrame(PWMH_mean_values)
    DWMH_mean_values_df = pd.DataFrame(DWMH_mean_values)

    # 设置 sub_dir 为 DataFrame 的索引，并命名为 'subject'
    PWMH_mean_values_df['subject'] = sub_dir
    DWMH_mean_values_df['subject'] = sub_dir

    # 调整列的顺序
    PWMH_columns = ['subject', 'Threshold'] + [col for col in PWMH_mean_values_df.columns if col not in ['subject', 'Threshold']]
    DWMH_columns = ['subject', 'Threshold'] + [col for col in DWMH_mean_values_df.columns if col not in ['subject', 'Threshold']]

    PWMH_mean_values_df = PWMH_mean_values_df[PWMH_columns]
    DWMH_mean_values_df = DWMH_mean_values_df[DWMH_columns]

    # 构建输出文件路径
    PWMH_output_file_path = os.path.join(shape_features_dir, f'{sub_dir}_average_PWMH_shape_features.xlsx')
    DWMH_output_file_path = os.path.join(shape_features_dir, f'{sub_dir}_average_DWMH_shape_features.xlsx')

    # 保存到Excel文件
    PWMH_mean_values_df.to_excel(PWMH_output_file_path, index=False)
    DWMH_mean_values_df.to_excel(DWMH_output_file_path, index=False)

    print(f'Saved mean values to {PWMH_output_file_path}')
    print(f'Saved mean values to {DWMH_output_file_path}')


#%%
# 初始化汇总的DataFrame
PWMH_combined_df = pd.DataFrame()
DWMH_combined_df = pd.DataFrame()

# 获取所有子目录
sub_dirs = [d for d in os.listdir(derivatives_dir) if
            os.path.isdir(os.path.join(derivatives_dir, d)) and d.startswith('sub-')]

# 遍历每个子目录
for sub_dir in sub_dirs:
    shape_features_dir = os.path.join(derivatives_dir, sub_dir, 'shape_features')

    # 构建文件路径
    PWMH_output_file_path = os.path.join(shape_features_dir, f'{sub_dir}_average_PWMH_shape_features.xlsx')
    DWMH_output_file_path = os.path.join(shape_features_dir, f'{sub_dir}_average_DWMH_shape_features.xlsx')

    # 读取数据
    if os.path.exists(PWMH_output_file_path):
        PWMH_data = pd.read_excel(PWMH_output_file_path, index_col='subject')
        PWMH_data.fillna('NA', inplace=True)  # 用 'NA' 替换空值
        PWMH_combined_df = PWMH_combined_df.append(PWMH_data)

    if os.path.exists(DWMH_output_file_path):
        DWMH_data = pd.read_excel(DWMH_output_file_path, index_col='subject')
        DWMH_data.fillna('NA', inplace=True)  # 用 'NA' 替换空值
        DWMH_combined_df = DWMH_combined_df.append(DWMH_data)

# 构建汇总文件的输出路径
PWMH_combined_file_path = os.path.join(derivatives_dir, 'PWMH_shape_features_per_subject.xlsx')
DWMH_combined_file_path = os.path.join(derivatives_dir, 'DWMH_shape_features_per_subject.xlsx')

# 保存汇总数据到Excel文件
PWMH_combined_df.to_excel(PWMH_combined_file_path)
DWMH_combined_df.to_excel(DWMH_combined_file_path)

print(f'Saved combined PWMH data to {PWMH_combined_file_path}')
print(f'Saved combined DWMH data to {DWMH_combined_file_path}')

#%%
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

def process_and_color_excel(file_path, output_file_path):
    # 读取Excel文件
    excel_data = pd.read_excel(file_path)

    # 将数据透视成每个受试者一行
    pivoted_data = excel_data.pivot(index='subject', columns='Threshold').reset_index()

    # 展平多层列名
    pivoted_data.columns = ['_'.join(map(str, col)).rstrip('_') for col in pivoted_data.columns.values]

    # 从原始数据框中提取有序特征列表
    ordered_features = [col for col in excel_data.columns if col not in ['subject', 'Threshold']]
    ordered_features = list(dict.fromkeys(ordered_features))  # 去除重复项，保持顺序

    # 创建新的DataFrame以保存重组数据
    reorganized_data = pd.DataFrame(pivoted_data['subject'])

    # 遍历每个阈值，然后是每个特征，以保持顺序
    thresholds = sorted(excel_data['Threshold'].unique())
    for threshold in thresholds:
        for feature in ordered_features:
            column_name = f"{feature}_{threshold}"
            if column_name in pivoted_data.columns:
                reorganized_data[f"{feature}_{threshold} voxels"] = pivoted_data[column_name]

    # 用'NA'替换缺失值
    reorganized_data.fillna('NA', inplace=True)

    # 保存重组数据到新的Excel文件
    reorganized_data.to_excel(output_file_path, index=False)

    # 使用openpyxl打开文件并添加颜色
    wb = openpyxl.load_workbook(output_file_path)
    ws = wb.active

    # 定义不同的背景颜色
    colors = ['FFCCCC', 'FF9999', 'FF6666', 'FF0000', 'CC0000', '990000']

    # 计算每个阈值的列范围并给这些列添加颜色
    col_start = 2
    for i, threshold in enumerate(thresholds):
        col_end = col_start + len(ordered_features) - 1
        fill = PatternFill(start_color=colors[i % len(colors)], end_color=colors[i % len(colors)], fill_type="solid")
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=col_start, max_col=col_end):
            for cell in col:
                cell.fill = fill
        col_start = col_end + 1

    # 保存修改后的文件
    wb.save(output_file_path)

# 对两个文件分别调用函数
process_and_color_excel(PWMH_combined_file_path,
                        os.path.join(derivatives_dir, 'PWMH_shape_features_per_subject_horizontal.xlsx'))

process_and_color_excel(DWMH_combined_file_path,
                        os.path.join(derivatives_dir, 'DWMH_shape_features_per_subject_horizontal.xlsx'))